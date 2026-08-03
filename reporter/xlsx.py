"""자산관리대장 엑셀 — 실물 양식 재현.

원본 양식(03. 정보자산관리지침 별첨 클라우드 보안인증(SaaS) 정보자산목록.xlsx)의
시트 구성·컬럼 순서·머리글 계층을 그대로 따르고, 계약 v1.0 레코드를 그 칸에 넣는다.
어떤 필드가 어느 칸으로 가는지는 config/sheet_map.yaml 에 있고 여기는 조립만 한다.

빈칸 표기가 이 파일의 두 번째 일이다. 값이 없는 이유가 다르면 조치도 다른데,
전부 공백으로 두면 그게 안 보인다. 수집기가 실어 보낸 사유를 글자로 옮긴다.

    -          이 자산유형에 개념 없음
    미식별      태그가 없다 (조치: 태그 부착)
    미설정      AWS 설정이 없다 (사실 자체가 등급 근거)
    API 미제공  API가 값을 주지 않는다
    미확인      권한 부족·조회 실패 (조치: 권한 부여 후 재수집. 자산 부재 아님)
    범위 밖      수집기가 아직 그 API를 안 부른다
    수기         AWS로는 얻을 수 없는 칸
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from advisor import classify, manual as manual_input, resolve

MARK = {
    resolve.NOT_APPLICABLE: "-",
    resolve.MISSING: "미식별",
    resolve.NOT_SET: "미설정",
    resolve.API_NULL: "API 미제공",
    resolve.UNVERIFIED: "미확인",
    resolve.OUT_OF_SCOPE: "범위 밖",
}
MANUAL = "수기"
GAP_MARKS = {"미식별", "미확인", "범위 밖", "판정불가"}

FORM_FILL = PatternFill("solid", fgColor="D9E1F2")      # 원본 양식 칸
EXTRA_FILL = PatternFill("solid", fgColor="E2EFDA")     # 인증기준이 더 요구하는 칸
GAP_FILL = PatternFill("solid", fgColor="FCE4E4")
LABEL_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

START_COL = 2  # 원본 양식이 B열부터 시작한다
WIDTH_SAMPLE = 200  # 열 너비를 재는 데 쓰는 행 수. 수천 행을 다 재면 느리다


# ── 여러 필드를 한 칸에 넣는 조립 함수들 ────────────────────────────────
# 계약은 값을 나눠 담고 표기 조립은 출력기 책임으로 둔다.

def _fn_extra_ips(asset):
    public = resolve.value_of(asset, "ip_public")
    return f"{public} (공인)" if public else "-"


def _fn_location(asset):
    region = resolve.value_of(asset, "region")
    if not region:
        return MARK[resolve.MISSING]
    az = resolve.value_of(asset, "az")
    return f"AWS {region}" + (f" / {az}" if az else "")


def _fn_dbms_name(asset):
    engine = resolve.value_of(asset, "engine")
    if not engine:
        return MARK[resolve.status(asset, "engine")]
    version = resolve.value_of(asset, "version")
    return f"{engine} {version}" if version else engine


def _fn_db_host(asset):
    # 양식은 온프레미스를 전제로 "설치 서버 호스트명"을 묻는다.
    # 관리형 데이터베이스는 그런 서버가 없다. 빈칸으로 두면 미식별로 오인된다.
    return "관리형 서비스 (설치 서버 없음)"


def _fn_db_address(asset):
    endpoint = resolve.value_of(asset, "endpoint")
    if not endpoint:
        return MARK[resolve.status(asset, "endpoint")]
    port = resolve.value_of(asset, "port")
    return f"{endpoint}:{port}" if port else endpoint


def _fn_resource_ref(asset):
    serial = resolve.value_of(asset, "serial_no") or ""
    return f"{asset['resource_type']} / {serial}"


def _fn_created_date(asset):
    value = resolve.value_of(asset, "created_at")
    return value[:10] if value else MARK[resolve.status(asset, "created_at")]


FUNCS = {
    "extra_ips": _fn_extra_ips,
    "location": _fn_location,
    "dbms_name": _fn_dbms_name,
    "db_host": _fn_db_host,
    "db_address": _fn_db_address,
    "resource_ref": _fn_resource_ref,
    "created_date": _fn_created_date,
}


def cell_value(asset, column, seq, manual=None):
    source = column["from"]
    if source == "seq":
        return seq
    if source == "manual":
        # 사람이 채운 값이 있으면 그걸 쓰고, 없으면 채워야 할 칸이라고 표시한다
        filled = manual_input.value_of(manual, asset["asset_id"], column["label"]) if manual else None
        return filled or MANUAL
    if source.startswith("const:"):
        return source.split(":", 1)[1]
    if source.startswith("fn:"):
        return FUNCS[source.split(":", 1)[1]](asset)
    if source.startswith("grade:"):
        return _grade_value(asset, source.split(":", 1)[1])

    state = resolve.status(asset, source)
    if state != resolve.PRESENT:
        return MARK[state]
    value = resolve.value_of(asset, source)
    if isinstance(value, bool):
        return "Y" if value else "N"
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    return value


def _grade_value(asset, axis):
    proposed = asset.get("grade_proposed")
    if axis == "confirmed":
        return "미확정"  # 도구는 이 칸을 절대 쓰지 않는다
    if not proposed:
        return "판정 전"
    if axis == "overall":
        return proposed["overall"] or "판정불가"
    return proposed[axis]["level"] or "판정불가"


# ── 시트 그리기 ────────────────────────────────────────────────────────

def _write(sheet, row, col, value, fill=None, bold=False, center=False):
    cell = sheet.cell(row=row, column=col, value=value)
    cell.border = BOX
    if fill:
        cell.fill = fill
    if bold:
        cell.font = Font(bold=True)
    if center:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def _cover(book, snapshot, gaps):
    meta = snapshot["meta"]
    sheet = book.active
    sheet.title = "표지"
    sheet["B3"] = "자산관리대장"
    sheet["B3"].font = Font(size=20, bold=True)
    sheet.merge_cells("B3:G3")
    sheet["B3"].alignment = Alignment(horizontal="center")

    uncovered = gaps["type1_manual"]["uncovered_types"]
    rows = [
        ("사업자명/소속", ""),
        ("문서등급", "대외비"),
        ("", ""),
        ("작성 근거", "정보자산관리지침 제6조② (별첨 정보자산목록 서식)"),
        ("작성 방식", "AWS API 자동 수집 + 태그 기반 조직정보 + 등급 제안"),
        ("스캔 실행", meta["run_id"]),
        ("대상 계정", meta["account_id"]),
        ("대상 리전", ", ".join(meta.get("regions", []))),
        ("수집 시각", meta.get("collected_at", "")),
        ("계약 버전", meta.get("contract_version", "")),
        ("식별 자산", f"{gaps['total_assets']}건"),
        ("갱신 주기", "반기 1회 (정보자산관리지침 제6조③)"),
        ("", ""),
        ("이 대장의 한계", "개인정보 보유 여부는 PII 태그가 유일한 근거이며 데이터 스캔은 하지 않음"),
        ("", "보안등급은 제안이며 확정은 관리책임자의 검토를 거침"),
        ("", f"수집기가 없는 자산유형 {len(uncovered)}종({', '.join(uncovered)})은 "
             "0건이 아니라 확인하지 않은 것임"),
    ]
    for i, (label, value) in enumerate(rows, start=9):
        if label:
            _write(sheet, i, 2, label, fill=LABEL_FILL, bold=True)
        sheet.cell(row=i, column=3, value=value)
    sheet.column_dimensions["B"].width = 18
    for col in "DEFG":
        sheet.column_dimensions[col].width = 16
    sheet.column_dimensions["C"].width = 78


def _ledger_sheet(book, spec, assets, extra_columns, manual=None):
    sheet = book.create_sheet(spec["name"])
    columns = spec["columns"] + extra_columns
    width = len(columns)

    sheet.cell(row=2, column=START_COL, value=spec["title"]).font = Font(size=14, bold=True)
    sheet.merge_cells(start_row=2, start_column=START_COL,
                      end_row=2, end_column=START_COL + width - 1)

    _write(sheet, 4, START_COL, "자산분류", fill=LABEL_FILL, bold=True)
    _write(sheet, 4, START_COL + 1, spec["asset_class"])
    _write(sheet, 5, START_COL, "Description", fill=LABEL_FILL, bold=True)
    _write(sheet, 5, START_COL + 1, spec["description"])
    for row in (4, 5):
        sheet.merge_cells(start_row=row, start_column=START_COL + 1,
                          end_row=row, end_column=START_COL + 5)

    # 8행 = 묶음 머리글, 9행 = 컬럼 머리글 (원본 양식과 같은 2단 구조)
    col = START_COL
    groups = dict(spec["groups"])
    groups["인증기준 추가 요구"] = len(extra_columns)
    for name, span in groups.items():
        fill = EXTRA_FILL if name == "인증기준 추가 요구" else FORM_FILL
        _write(sheet, 8, col, name, fill=fill, bold=True, center=True)
        if span > 1:
            sheet.merge_cells(start_row=8, start_column=col, end_row=8, end_column=col + span - 1)
        col += span

    form_count = len(spec["columns"])
    for i, column in enumerate(columns):
        fill = FORM_FILL if i < form_count else EXTRA_FILL
        _write(sheet, 9, START_COL + i, column["label"], fill=fill, bold=True, center=True)

    if not assets:
        _write(sheet, 10, START_COL, "이번 스캔 범위에서 0건")
        sheet.merge_cells(start_row=10, start_column=START_COL,
                          end_row=10, end_column=START_COL + width - 1)
    for seq, asset in enumerate(assets, 1):
        for i, column in enumerate(columns):
            value = cell_value(asset, column, seq, manual)
            cell = _write(sheet, 9 + seq, START_COL + i, value)
            if value in GAP_MARKS:
                cell.fill = GAP_FILL

    sheet.freeze_panes = f"{get_column_letter(START_COL + 3)}10"
    for i, column in enumerate(columns):
        values = [column["label"]] + [
            str(sheet.cell(row=9 + s, column=START_COL + i).value or "")
            for s in range(1, min(len(assets), WIDTH_SAMPLE) + 1)
        ]
        size = max(len(v) for v in values) + 3
        sheet.column_dimensions[get_column_letter(START_COL + i)].width = min(max(size, 10), 40)
    return sheet


def _summary_sheet(book, snapshot, sheet_map):
    """인증신청 양식 v2.5 '인증범위내 주요 자산' 요약표."""
    conf = sheet_map["summary_table"]
    sheet = book.create_sheet("인증신청서 요약표")

    sheet.cell(row=2, column=2, value="인증범위내 주요 자산").font = Font(size=14, bold=True)
    sheet.cell(row=3, column=2, value=conf["count_rule"]).alignment = Alignment(wrap_text=True)
    sheet.merge_cells("B3:F4")

    for i, label in enumerate(conf["columns"]):
        _write(sheet, 6, 2 + i, label, fill=FORM_FILL, bold=True, center=True)

    buckets = {}
    for asset in resolve.assets_of(snapshot):
        key = (
            resolve.value_of(asset, "service_name") or "서비스 미지정",
            conf["category_map"].get(resolve.value_of(asset, "asset_type"), "기타"),
            resolve.value_of(asset, "region") or "글로벌",
            resolve.value_of(asset, "owner_dept") or MARK[resolve.MISSING],
        )
        buckets[key] = buckets.get(key, 0) + 1

    row = 7
    for (service, division, region, dept), count in sorted(buckets.items()):
        for i, value in enumerate([service, division, f"{count}대", f"AWS {region}", dept]):
            cell = _write(sheet, row, 2 + i, value)
            if value in GAP_MARKS or value == "서비스 미지정":
                cell.fill = GAP_FILL
        row += 1

    row += 1
    sheet.cell(row=row, column=2, value="중복 제외 후보 (자동 제외하지 않음)").font = Font(bold=True)
    row += 1
    dupes = _duplicate_candidates(snapshot)
    if not dupes:
        sheet.cell(row=row, column=2, value="없음")
    for note in dupes:
        sheet.cell(row=row, column=2, value=note)
        row += 1

    for col, width in zip("BCDEF", (26, 16, 10, 20, 20)):
        sheet.column_dimensions[col].width = width


def _duplicate_candidates(snapshot):
    """서버 수 산정 시 중복 제외 후보.

    신청 양식은 네트워크 존과 OS 세부 버전·모델이 완전히 같을 때만 중복 제외를
    허용한다. 판단은 사람 몫이라 도구는 후보만 표시한다.
    """
    groups = {}
    for asset in snapshot["asset_types"].get("서버", {}).get("assets", []):
        key = (resolve.value_of(asset, "subnet_id"), resolve.value_of(asset, "os"),
               resolve.value_of(asset, "version"), resolve.value_of(asset, "model"))
        name = resolve.value_of(asset, "asset_name") or asset["asset_id"].rsplit("/", 1)[-1]
        groups.setdefault(key, []).append(name)

    notes = []
    for (subnet, os_name, version, model), names in groups.items():
        if len(names) > 1:
            notes.append(f"{', '.join(names[:6])} 은 서브넷({subnet})과 "
                         f"OS({os_name} {version}), 모델({model})이 동일")
    return notes


def _gap_sheet(book, gaps, unclassified):
    sheet = book.create_sheet("갭 리포트")
    for i, label in enumerate(["유형", "자산유형", "항목", "건수", "표본", "조치"]):
        _write(sheet, 1, 1 + i, label, fill=FORM_FILL, bold=True, center=True)

    def bucket(rows, kind, action):
        return [[kind, r["asset_type"], f"{r['item_name']} ({r['field']})",
                 r["count"], ", ".join(r["samples"]), action] for r in rows]

    rows = []
    rows += bucket(gaps["type2_missing_tag"], "② 태그 미입력", "태그 부착 후 재수집")
    rows += bucket(gaps["not_configured"], "미설정", "설정 자체가 없음. 등급 근거로 사용")
    rows += bucket(gaps["api_null"], "API 미제공", "AWS가 값을 주지 않음")
    rows += bucket(gaps["unverified"], "미확인", "권한 부여 후 재수집 (자산 부재 아님)")
    rows += bucket(gaps["out_of_scope"], "범위 밖", "수집기가 해당 API를 부르지 않음")

    for action in gaps["type1_manual"]["permanent"]:
        rows.append(["① 수기 등재 필요", action.get("owner", "-"), action["item_name"],
                     action.get("affected_assets", ""), "",
                     action.get("action", "수기 입력 시트에서 작성")])
    for type_name in gaps["type1_manual"]["uncovered_types"]:
        rows.append(["① 수집기 없음", type_name, "이 유형은 훑지 않았음", "", "",
                     "0건이 아니라 미확인. 별도 식별 필요"])
    for item in gaps["blocked"]:
        rows.append(["차단", "-", item["field"], item.get("affected_assets", ""),
                     ", ".join(x.rsplit("/", 1)[-1] for x in item.get("sample_asset_ids", [])[:3]),
                     f"{item['reason']} — 작업 목록과 섞지 말 것"])
    for issue in gaps["collection_issues"]:
        rows.append(["수집 실패", "-", str(issue), "", "", "이 범위는 0건을 신뢰하지 말 것"])
    for name, reason in unclassified:
        rows.append(["분류 보류", "-", name, "", "", reason])
    for item in gaps["type3_pending_confirm"][:50]:
        rows.append(["③ 확정 대기", item["asset_type"], item["asset_name"], "",
                     item["overall"] or "판정불가", "검토 후 확정등급 기입"])

    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            _write(sheet, r, c, value)

    sheet.freeze_panes = "A2"
    for col, width in zip("ABCDEF", (16, 20, 42, 8, 34, 40)):
        sheet.column_dimensions[col].width = width


def _manual_sheet(book, sheet_map, buckets, manual, common_labels):
    """AWS로 못 얻는 칸의 현황. 채운 값과 아직 빈 칸을 같이 보여준다."""
    sheet = book.create_sheet("수기 입력")
    for i, label in enumerate(["구분", "대상", "항목", "값", "작성자", "작성일"]):
        _write(sheet, 1, 1 + i, label, fill=EXTRA_FILL, bold=True, center=True)

    entries = (manual or {}).get("entries", {})

    def row_for(scope, about, label, kind):
        entry = entries.get(scope, {}).get(label) or {}
        return [kind, about, label, entry.get("value", ""),
                entry.get("by", ""), (entry.get("at") or "")[:10]]

    rows = [row_for(manual_input.COMMON, "대장 전체", label, "대장 공통")
            for label in common_labels]
    for sheet_name, label in manual_input.manual_columns(sheet_map):
        for asset in buckets.get(sheet_name, []):
            about = resolve.value_of(asset, "asset_name") or asset["asset_id"].rsplit("/", 1)[-1]
            rows.append(row_for(asset["asset_id"], about, label, f"{sheet_name} 시트"))

    for r, values in enumerate(rows, start=2):
        for c, value in enumerate(values, start=1):
            cell = _write(sheet, r, c, value)
            if c == 4 and not value:
                cell.fill = GAP_FILL

    sheet.freeze_panes = "A2"
    for col, width in zip("ABCDEF", (16, 26, 34, 34, 12, 12)):
        sheet.column_dimensions[col].width = width


def write(snapshot, gaps, path, sheet_map=None, buckets=None, manual=None,
          common_labels=()):
    sheet_map = sheet_map or classify.load_sheet_map()
    if buckets is None:
        buckets = classify.assign(resolve.assets_of(snapshot), sheet_map)
    unclassified = [(resolve.value_of(a, "asset_name") or a["asset_id"],
                     classify.unclassified_reason(a))
                    for a in buckets[classify.UNKNOWN_SHEET]]

    book = Workbook()
    _cover(book, snapshot, gaps)
    for spec in sheet_map["sheets"]:
        _ledger_sheet(book, spec, buckets[spec["name"]],
                      sheet_map["certification_columns"], manual)
    _summary_sheet(book, snapshot, sheet_map)
    _gap_sheet(book, gaps, unclassified)
    _manual_sheet(book, sheet_map, buckets, manual, common_labels)
    book.save(path)
    return path
