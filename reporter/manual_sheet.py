"""수기 입력 템플릿 xlsx.

담당자가 실제로 채워 넣을 빈 양식을 만든다.
**담당자가 백지에서 시작하지 않게 하는 것이 목적이다.**

reporter/ 는 원래 담당 B의 영역이지만 이 파일만 예외다.
판정이 아니라 수집 불가 항목 안내이기 때문이다.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CONFIG_DIR = Path("config")
OUTPUT_ROOT = Path("output")

HEADERS = [
    ("자산유형", 16),
    ("항목명", 30),
    ("예시", 34),
    ("왜 수기인가", 46),
    ("근거", 52),
    ("담당", 24),
    ("해야 할 일", 46),
    ("작성란", 40),
]

#: 설비·시설은 빈 칸으로 두면 그 자체가 결함이다. 문구를 미리 채워둔다.
PREFILLED = {
    "facility_equipment": (
        "책임공유모델상 CSP 책임영역. AWS Artifact ISO 27001·SOC 2 보고서 첨부"
    ),
    "facility_site": (
        "책임공유모델상 CSP 책임영역. AWS Artifact ISO 27001·SOC 2 보고서 첨부"
    ),
}

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")


def load_manual_items(config_dir: Path | None = None) -> dict:
    path = (config_dir or CONFIG_DIR) / "manual_items.yaml"
    return yaml.safe_load(path.read_text(encoding="utf-8"))["items"]


def build(
    *,
    payload: dict | None = None,
    config_dir: Path | None = None,
    out_path: Path | None = None,
) -> Path:
    """수기 입력 템플릿을 만든다.

    payload(assets.json)를 주면 태그로 해결되는 항목에 영향 자산 수를 함께 적는다.
    """
    items = load_manual_items(config_dir)
    todo = (payload or {}).get("manual_todo", {})
    affected = {
        row["key"]: row
        for rows in todo.get("by_owner", {}).values()
        for row in rows
    }

    workbook = Workbook()
    _sheet_permanent(workbook.active, items)
    _sheet_tag_fixable(workbook.create_sheet("태그로 해결"), items, affected)
    _sheet_guide(workbook.create_sheet("읽는 법"))

    out_path = out_path or (OUTPUT_ROOT / "수기입력_템플릿.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)
    return out_path


def _write_header(sheet: Any, headers: list[tuple[str, int]]) -> None:
    for col, (title, width) in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="center")
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.freeze_panes = "A2"


def _sheet_permanent(sheet: Any, items: dict) -> None:
    """AWS 밖 자산 — 영원히 수기다."""
    sheet.title = "수기 등재"
    _write_header(sheet, HEADERS)

    row = 2
    for key, meta in items.items():
        if meta.get("auto_after_fix", False):
            continue
        values = [
            meta.get("isms_asset_type", ""),
            meta.get("item_name", key),
            ", ".join(meta.get("examples") or []),
            _clean(meta.get("reason")),
            _clean(meta.get("evidence")),
            meta.get("owner", ""),
            _clean(meta.get("action")),
            PREFILLED.get(key, ""),
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == len(values):
                cell.fill = _INPUT_FILL
        row += 1


def _sheet_tag_fixable(sheet: Any, items: dict, affected: dict) -> None:
    """태그만 달면 다음 실행부터 자동 — 한 번 투자하면 끝나는 일이다."""
    headers = [
        ("항목명", 30), ("영향 자산 수", 14), ("비율", 10),
        ("담당", 26), ("해야 할 일", 60), ("샘플 자산", 52), ("완료 표시", 12),
    ]
    _write_header(sheet, headers)

    rows = []
    for key, meta in items.items():
        if not meta.get("auto_after_fix", False):
            continue
        hit = affected.get(key, {})
        rows.append((
            meta.get("item_name", key),
            hit.get("affected_assets", 0),
            hit.get("affected_ratio", "—"),
            meta.get("owner", ""),
            _clean(meta.get("action")),
            "\n".join(hit.get("sample_asset_ids", [])),
        ))

    # 영향 자산이 많은 것부터. "태그 하나로 62건 해결"이 우선순위 판단의 근거다.
    rows.sort(key=lambda r: -r[1])
    for index, values in enumerate(rows, start=2):
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=index, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row=index, column=len(headers)).fill = _INPUT_FILL


def _sheet_guide(sheet: Any) -> None:
    sheet.column_dimensions["A"].width = 110
    lines = [
        ("이 파일은 무엇인가", True),
        ("AWS API로 수집할 수 없는 항목을 담당자가 채워 넣기 위한 양식입니다.", False),
        ("자동 수집분은 별도 JSON(assets.json)에 있고, 이 파일과 합쳐야 자산관리대장이 완성됩니다.", False),
        ("", False),
        ("시트가 둘로 나뉜 이유", True),
        ("[수기 등재] AWS 밖 자산입니다. 매번 손으로 관리해야 합니다.", False),
        ("[태그로 해결] 태그를 한 번 달면 다음 실행부터 자동으로 채워집니다.", False),
        ("둘은 성격이 완전히 다른 일이라 섞지 않았습니다.", False),
        ("태그 작업을 먼저 하시면 손으로 채울 칸이 줄어듭니다.", False),
        ("", False),
        ("설비·시설 칸이 미리 채워져 있는 이유", True),
        ("클라우드에서는 CSP 책임영역이라 조직이 관리할 자산이 없습니다.", False),
        ("다만 사유 없이 비워두면 그 자체가 심사 결함이 되므로 제외 사유를 미리 적어두었습니다.", False),
        ("AWS Artifact에서 ISO 27001·SOC 2 보고서를 받아 함께 보관하십시오.", False),
        ("", False),
        ("주의", True),
        ("이 도구는 적합·부적합을 판정하지 않습니다. 미확인·미입력 사실만 제시합니다.", False),
        ("보안등급 확정은 사람의 몫입니다.", False),
    ]
    for index, (text, is_heading) in enumerate(lines, start=1):
        cell = sheet.cell(row=index, column=1, value=text)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if is_heading:
            cell.font = Font(bold=True)
            cell.fill = _SECTION_FILL


def _clean(text: Any) -> str:
    if not text:
        return ""
    return " ".join(str(text).split())
