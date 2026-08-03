"""S3 검증 — 보안시스템 수집기 + 수집 불가 항목 + 담당자 작업 지시."""

from __future__ import annotations

from pathlib import Path

import boto3
import pytest
from moto import mock_aws
from openpyxl import load_workbook

from collector import reasons as R
from collector.collect import run_collect
from collector.extract import run_extract
from collector.manual import load_manual_items
from collector.registry import discover
from reporter import manual_sheet
from tests.aws_fixture import REGION, build_environment

CONFIG_DIR = Path(__file__).resolve().parent.parent / "config"


@pytest.fixture(autouse=True)
def fake_credentials(monkeypatch):
    for key, value in {
        "AWS_ACCESS_KEY_ID": "testing",
        "AWS_SECRET_ACCESS_KEY": "testing",
        "AWS_SECURITY_TOKEN": "testing",
        "AWS_SESSION_TOKEN": "testing",
        "AWS_DEFAULT_REGION": REGION,
    }.items():
        monkeypatch.setenv(key, value)


@pytest.fixture
def payload(tmp_path):
    """침입탐지(GuardDuty)는 있고 침입차단(WAF)·침입방지는 없는 환경.

    보안시스템이 통째로 비어 있으면 결함사례 1이지만, 일부만 있는 상태가
    더 위험하다. "보안시스템 5건 있음"으로 뭉개지는지 확인해야 한다.
    """
    raw_root = tmp_path / "raw"

    @mock_aws
    def _build():
        build_environment(REGION)
        boto3.client("guardduty", region_name=REGION).create_detector(Enable=True)
        boto3.client("kms", region_name=REGION).create_key(Description="prd 암호화 키")
        boto3.client("secretsmanager", region_name=REGION).create_secret(
            Name="prd/db/password", SecretString="x")
        return run_collect(regions=(REGION,), root=raw_root)

    result = _build()
    return run_extract(
        run_id=result["run_id"], raw_root=raw_root,
        out_root=tmp_path / "norm", config_dir=CONFIG_DIR,
    )


def rows_of(payload, asset_type):
    return payload["asset_types"][asset_type]["manual_required"]


def find_row(payload, asset_type, item_name):
    for row in rows_of(payload, asset_type):
        if row["item_name"] == item_name:
            return row
    raise AssertionError(f"{asset_type}에 '{item_name}' 행이 없다")


# --------------------------------------------------------------------------
# 보안시스템 수집기
# --------------------------------------------------------------------------

def test_보안시스템_수집기가_등록된다():
    names = {type(c).__name__ for c in discover()}
    assert "SecurityCollector" in names


def test_보안시스템_수집기도_읽기_전용이다():
    for collector in discover():
        assert type(collector).write_actions() == [], type(collector).__name__


def test_활성화된_보안서비스가_자산으로_잡힌다(payload):
    assets = payload["asset_types"]["보안시스템"]["assets"]
    kinds = {a["resource_type"] for a in assets}
    assert "guardduty_detector" in kinds
    assert "kms_key" in kinds
    assert "secretsmanager_secret" in kinds


def test_AWS관리형_KMS키는_자산에서_제외된다(payload):
    """AWS 관리형 키는 계정 부속이지 조직이 관리하는 자산이 아니다."""
    keys = [a for a in payload["asset_types"]["보안시스템"]["assets"]
            if a["resource_type"] == "kms_key"]
    assert keys, "고객관리형 키가 하나는 잡혀야 한다"
    assert all(a["usage"]["value"] == "암호키(고객관리형)" for a in keys)


def test_태그API를_안_부른_자산은_TAG_ABSENT가_아니라_OUT_OF_SCOPE(payload):
    """kms:ListResourceTags를 안 불렀는데 '태그를 다세요'라고 하면 틀린 안내다."""
    key = next(a for a in payload["asset_types"]["보안시스템"]["assets"]
               if a["resource_type"] == "kms_key")
    assert key["owner_dept"]["reason"] == R.OUT_OF_SCOPE
    assert "ListResourceTags" in key["owner_dept"]["hint"]


def test_미활성_서비스는_NOT_CONFIGURED로_기록된다(tmp_path):
    """Security Hub·Shield는 미활성 시 예외를 던진다. 그게 '없다'는 증적이다."""
    raw_root = tmp_path / "raw"

    @mock_aws
    def _build():
        return run_collect(regions=(REGION,), root=raw_root)

    result = _build()
    import json
    dump = json.loads(
        (raw_root / result["run_id"] / result["account_id"] / REGION / "security.json")
        .read_text(encoding="utf-8"))
    assert dump["data"]["describe_hub"]["__status__"] == R.NOT_CONFIGURED
    assert dump["data"]["describe_subscription"]["__status__"] == R.NOT_CONFIGURED


# --------------------------------------------------------------------------
# 결함사례 1 — 있는 것과 없는 것을 갈라야 한다
# --------------------------------------------------------------------------

def test_존재하는_통제는_요구항목에서_빠진다(payload):
    """GuardDuty가 있으므로 침입탐지시스템은 미충족이 아니다."""
    names = {row["item_name"] for row in rows_of(payload, "보안시스템")}
    assert "침입탐지시스템" not in names


def test_없는_통제는_0건으로_드러난다(payload):
    """WAF가 없으므로 침입차단시스템은 0건이다.

    보안시스템 유형에 자산이 5건 있다고 해서 침입차단이 있는 게 아니다.
    이걸 뭉개면 결함사례 1을 잡으라고 만든 기능이 오히려 결함을 가린다.
    """
    assert payload["asset_types"]["보안시스템"]["asset_count"] > 0

    firewall = find_row(payload, "보안시스템", "침입차단시스템")
    assert firewall["collected_count"] == 0
    assert firewall["auto_after_fix"] is False
    assert firewall["owner"] and firewall["action"], "누가 무엇을 해야 하는지가 없다"
    assert firewall["note"] == "AWS 수집 결과 0건"

    dlp = find_row(payload, "보안시스템", "개인정보유출방지시스템")
    assert dlp["collected_count"] == 0
    assert "결함사례 1" in dlp["evidence"]
    assert "DRM" in " ".join(dlp["examples"])


def test_수기_등재_건수를_AWS_수집_건수로_적지_않는다(payload):
    """'내부정보 유출통제 시스템 5건 등재'는 거짓이다. 그 5건은 GuardDuty·KMS다."""
    rows = [r for rows in payload["manual_todo"]["by_owner"].values() for r in rows]
    dlp = next(r for r in rows if r["key"] == "security_system_dlp")
    assert dlp["currently_registered"] == 0


def test_설비와_시설은_수기_전용으로_남는다(payload):
    for type_name in ("설비", "시설"):
        block = payload["asset_types"][type_name]
        assert block["asset_count"] == 0
        assert block["manual_required"], f"{type_name}에 수기 지시가 없다"
        assert all(row["auto_after_fix"] is False for row in block["manual_required"])


# --------------------------------------------------------------------------
# 담당자 작업 지시
# --------------------------------------------------------------------------

def test_태그로_해결되는_것과_영구_수기가_섞이지_않는다(payload):
    todo = payload["manual_todo"]
    rows = [r for rows in todo["by_owner"].values() for r in rows]
    auto = [r for r in rows if r["auto_after_fix"]]
    permanent = [r for r in rows if not r["auto_after_fix"]]

    assert auto and permanent
    assert todo["summary"]["auto_after_fix"] == len(auto)
    assert todo["summary"]["permanent"] == len(permanent)
    # 태그 항목에는 영향 자산 수가 붙어야 우선순위를 정할 수 있다
    assert all(r["affected_assets"] > 0 for r in auto)
    assert all("affected_ratio" in r for r in auto)


def test_영향_자산_수와_샘플이_붙는다(payload):
    rows = [r for rows in payload["manual_todo"]["by_owner"].values() for r in rows]
    owner = next(r for r in rows if r["key"] == "asset_owner")
    assert owner["affected_assets"] > 0
    assert owner["sample_asset_ids"]
    assert owner["affected_ratio"].endswith("%")


def test_조건부_필수는_해당_자산만_센다(payload):
    """HandlePI=Y인 자산만 PIItems가 필수다.

    조건을 안 보면 개인정보를 안 다루는 자산까지 세어 숫자가 부풀고
    우선순위가 뒤집힌다.
    """
    rows = [r for rows in payload["manual_todo"]["by_owner"].values() for r in rows]
    items = next((r for r in rows if r["key"] == "personal_info_items"), None)
    if items is None:
        pytest.skip("HandlePI=Y 자산이 없는 환경")

    pi_assets = [
        a for block in payload["asset_types"].values() for a in block["assets"]
        if a.get("has_personal_info", {}).get("value") is True
    ]
    assert items["affected_assets"] <= len(pi_assets)


def test_권한부족과_조회실패는_따로_보고된다(payload):
    """권한 문제를 자산 문제로 오인하면 대장 전체가 틀린다."""
    blocked = payload["manual_todo"]["blocked"]
    assert blocked
    backup = next(b for b in blocked if b["field"] == "infra_facts.backup_exists")
    assert backup["reason"] in R.NOT_ABSENCE_REASONS
    assert "자산이 없다는 뜻이 아니다" in backup["note"]

    # 작업 목록에는 안 섞인다
    rows = [r for rows in payload["manual_todo"]["by_owner"].values() for r in rows]
    assert all("backup_exists" not in r.get("action", "") for r in rows)


def test_모든_수기항목이_누가_왜_어떻게를_갖는다():
    """빈 칸으로 두지 않는다는 원칙을 설정 파일 단계에서 강제한다."""
    for key, meta in load_manual_items(CONFIG_DIR).items():
        assert meta.get("owner"), f"{key}: 담당이 없다"
        assert meta.get("action"), f"{key}: 해야 할 일이 없다"
        assert meta.get("reason"), f"{key}: 왜 수기인지가 없다"
        assert meta.get("evidence"), f"{key}: 근거가 없다"
        assert "auto_after_fix" in meta, f"{key}: 태그로 해결되는지 구분이 없다"


# --------------------------------------------------------------------------
# 수기 입력 템플릿
# --------------------------------------------------------------------------

def test_수기_템플릿이_만들어진다(payload, tmp_path):
    out = manual_sheet.build(
        payload=payload, config_dir=CONFIG_DIR, out_path=tmp_path / "template.xlsx")
    assert out.exists()

    workbook = load_workbook(out)
    assert workbook.sheetnames == ["수기 등재", "태그로 해결", "읽는 법"]

    sheet = workbook["수기 등재"]
    headers = [c.value for c in sheet[1]]
    assert headers == ["자산유형", "항목명", "예시", "왜 수기인가", "근거", "담당", "해야 할 일", "작성란"]
    assert sheet.max_row > 1


def test_설비_시설은_제외사유가_미리_채워진다(payload, tmp_path):
    """사유 없이 비워두면 그 자체가 결함이다."""
    out = manual_sheet.build(
        payload=payload, config_dir=CONFIG_DIR, out_path=tmp_path / "t.xlsx")
    sheet = load_workbook(out)["수기 등재"]

    prefilled = {r[1]: r[7] for r in sheet.iter_rows(min_row=2, values_only=True) if r[7]}
    assert len(prefilled) == 2
    for text in prefilled.values():
        assert "책임공유모델" in text
        assert "Artifact" in text


def test_템플릿의_태그시트가_영향_자산_수로_정렬된다(payload, tmp_path):
    """'태그 하나로 62건 해결'이 우선순위 판단의 근거다."""
    out = manual_sheet.build(
        payload=payload, config_dir=CONFIG_DIR, out_path=tmp_path / "t.xlsx")
    sheet = load_workbook(out)["태그로 해결"]

    counts = [r[1] for r in sheet.iter_rows(min_row=2, values_only=True)]
    assert counts == sorted(counts, reverse=True)


def test_템플릿은_assets_json_없이도_만들어진다(tmp_path):
    """수집을 아직 안 돌린 담당자도 양식을 먼저 받을 수 있어야 한다."""
    out = manual_sheet.build(config_dir=CONFIG_DIR, out_path=tmp_path / "t.xlsx")
    sheet = load_workbook(out)["태그로 해결"]
    assert sheet.max_row > 1


def test_요구항목마다_누가_어떻게가_붙는다(payload):
    """'담당: 미지정 / 조치: 수기 확인 필요'는 이 도구가 없애려던 빈 칸이다."""
    for type_name, block in payload["asset_types"].items():
        for row in block["manual_required"]:
            where = f"{type_name}.{row['item_name']}"
            assert row["owner"] != "미지정", where
            assert row["action"] != "수기 확인 필요", where
            assert row["evidence"], where


def test_필드_요구항목은_manual_ref_없이도_안내를_찾는다(payload):
    """required_items에 manual_ref를 손으로 안 달아도 fields로 연결된다."""
    row = find_row(payload, "데이터(DBMS)", "관리 부서명")
    assert row["auto_after_fix"] is True
    assert "OwnerDept" in row["action"]
    assert row["manual_ref"] is None
